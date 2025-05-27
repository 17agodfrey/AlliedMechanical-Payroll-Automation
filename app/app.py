import pandas as pd
import re
import os
from formatting import apply_formatting_to_excel
from taxes import calculate_tax, get_tax
from openpyxl.utils import get_column_letter


def update_bonus_vacation(cell_value, total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i):
    if isinstance(cell_value, str):
        if 'BN' in cell_value:
            match = re.search(r'BN\s*([\d,]+\.\d{2})', cell_value)
            if match:
                total_bonus += float(match.group(1).replace(',', ''))
                employee_has_BN = True
                
        elif 'VAC' in cell_value:
            match = re.search(r'VAC\s*([\d,]+\.\d{2})', cell_value)
            if match:
                total_vacation += float(match.group(1).replace(',', ''))
                employee_has_VAC = True
                
        #handle special case: VAC and BN both reported for a single employee. 
        # WARNING: this only works if the rows are right next to each other. Ie. 3,4 or 1,2... etc. 
        if (employee_has_VAC and employee_has_BN):
            c_employee_row_i = c_employee_row_i - 1       
             
    return total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i

def process_payroll_file(filepath, tax_data):
    df = pd.read_excel(filepath, sheet_name='Payroll Register', header=None)
    job_lookup = {}
    current_employee = None
    c_employee_row_i = 0;

    total_bonus = 0.0
    total_vacation = 0.0
    un_coded_pay = 0.0
    gross_total = 0.0
    
    errors = []
    employee_has_BN = False
    employee_has_VAC = False

    for i in range(len(df)):
        row = df.iloc[i]                

        if not isinstance(row[0], str):
            continue

        if "Associate ID" in row[0]:
            current_employee = row[0].split('\n')[0] if '\n' in row[0] else row[0]
            current_employee = re.sub(r'\s*,\s*', ', ', str(current_employee).strip()) ## Normalize name, just in case 
            c_employee_row_i = 1
            employee_has_BN = False
            employee_has_VAC = False
            
            ##handle bonus and vacation
            total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i = update_bonus_vacation(row[6], total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i)
            
            job_number = '0001'
            if "W-In Cost" in row[0]:
                match = re.search(r'W-In Cost:\s*(\d{4,})-(\d{4,})', row[0])
                job_number = match.group(2) if match else '0001'

            try:
                reg_pay = float(str(row[4]).replace(',', '')) if pd.notna(row[4]) else 0.0
                ot_pay = float(str(row[5]).replace(',', '')) if pd.notna(row[5]) else 0.0
                if reg_pay == 0.0 and ot_pay == 0.0:
                    if isinstance(row[3], str) and 'UTO' in row[3]:
                        c_employee_row_i = c_employee_row_i - 1
                    continue

                emp_tax, memo_401k, errors_get_tax = get_tax(tax_data, current_employee, c_employee_row_i)
                if pd.isna(memo_401k): memo_401k = 0.0
                
                if errors_get_tax:
                    errors.extend(errors_get_tax)

                key = (current_employee, job_number)
                if key in job_lookup:
                    job_lookup[key]['Reg Pay'] += reg_pay
                    job_lookup[key]['OT Pay'] += ot_pay
                    job_lookup[key]['Emp Tax'] += emp_tax
                    job_lookup[key]['Memo 401k'] += memo_401k
                else:
                    job_lookup[key] = {
                        'Employee': current_employee,
                        'Job Number': job_number,
                        'Reg Pay': reg_pay,
                        'OT Pay': ot_pay,
                        'Emp Tax': emp_tax,
                        'Memo 401k': memo_401k
                    }
            except Exception as e:
                errors.append(f"Error parsing Associate ID row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee}. ERROR : {e}")
                # print(f"Error parsing Associate ID row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee}. ERROR : {e}")

        elif "W-In Cost" in row[0] and current_employee:
            match = re.search(r'W-In Cost:\s*(\d{4,})-(\d{4,})', row[0])
            job_number = match.group(2) if match else '0001'
            
            c_employee_row_i = c_employee_row_i + 1

            total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i = update_bonus_vacation(row[6], total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i)
            
            try:
                reg_pay = float(str(row[4]).replace(',', '')) if pd.notna(row[4]) else 0.0
                ot_pay = float(str(row[5]).replace(',', '')) if pd.notna(row[5]) else 0.0
                if reg_pay == 0.0 and ot_pay == 0.0:
                    continue
                
                emp_tax, memo_401k, errors_get_tax = get_tax(tax_data, current_employee, c_employee_row_i)
                if pd.isna(memo_401k): memo_401k = 0.0

                
                if errors_get_tax:
                    errors.extend(errors_get_tax)

                key = (current_employee, job_number)
                if key in job_lookup:
                    job_lookup[key]['Reg Pay'] += reg_pay
                    job_lookup[key]['OT Pay'] += ot_pay
                    job_lookup[key]['Emp Tax'] += emp_tax
                    job_lookup[key]['Memo 401k'] += memo_401k
                else:
                    job_lookup[key] = {
                        'Employee': current_employee,
                        'Job Number': job_number,
                        'Reg Pay': reg_pay,
                        'OT Pay': ot_pay,
                        'Emp Tax': emp_tax,
                        'Memo 401k': memo_401k
                    }
            except Exception as e:
                errors.append(f"Error parsing row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee}. ERROR : {e}")
                # print(f"Error parsing row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee} : {e}")
                
        #handle UTO
        elif isinstance(row[3], str) and 'UTO' in row[3]:
                continue
        # Handle rows with no W-In Cost but containing Reg/OT pay (Uncoded pay, but not bonues and etc) 
        elif "H Dept" in row[0] and current_employee:              
            c_employee_row_i = c_employee_row_i + 1
            
            total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i = update_bonus_vacation(row[6], total_bonus, total_vacation, employee_has_BN, employee_has_VAC, c_employee_row_i)
            
            try:
                reg_pay = float(str(row[4]).replace(',', '')) if pd.notna(row[4]) else 0.0
                ot_pay = float(str(row[5]).replace(',', '')) if pd.notna(row[5]) else 0.0
                if reg_pay > 0.0 or ot_pay > 0.0:
                    un_coded_pay += reg_pay + ot_pay
            except Exception as e:
                errors.append((f"Error parsing Uncoded Pay row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee}. ERROR : {e}"))
                # print(f"Error parsing Uncoded Pay row ({c_employee_row_i}) in Payroll file for Current Employee: {current_employee}. ERROR : {e}")

        elif isinstance(row[7], str) and 'Gross' in row[7]:
            match = re.search(r'Gross\s*([\d,]+\.\d{2})', row[7])
            if match:
                try:
                    gross_total += float(match.group(1).replace(',', ''))
                except Exception as e:
                    errors.append(f"Error parsing Gross Pay in Payroll file. ERROR : {e}")
                    # print(f"Error parsing Gross Pay in Payroll file. ERROR : {e}")


    ######################################################################## 

    job_data = list(job_lookup.values())

    df_jobs = pd.DataFrame(job_data)
    job_numbers = sorted(df_jobs['Job Number'].unique())

    # Build column multi-index: (Job, Reg/O/T)
    columns = []
    for job in job_numbers:
        columns.append((job, 'Reg'))
        columns.append((job, 'O/T'))
        columns.append((job, 'Emp Tax'))
        columns.append((job, 'Memo 401k'))  # Add Memo 401k column if needed

    multi_columns = pd.MultiIndex.from_tuples(columns)

    # Build data rows for each employee
    employee_rows = []
    employee_names = []
    for employee in df_jobs['Employee'].unique():
        emp_jobs = df_jobs[df_jobs['Employee'] == employee]
        row = []
        for job in job_numbers:
            job_entry = emp_jobs[emp_jobs['Job Number'] == job]
            if not job_entry.empty:
                row.append(job_entry['Reg Pay'].values[0])
                row.append(job_entry['OT Pay'].values[0])
                row.append(job_entry['Emp Tax'].values[0])
                row.append(job_entry['Memo 401k'].values[0])
            else:
                row.extend([None, None, None, None])
        employee_rows.append(row)
        employee_names.append(employee)

    output_df = pd.DataFrame(employee_rows, columns=multi_columns)
    output_df.insert(0, ('Job Number', ''), employee_names)

    # Build header rows
    file_title = os.path.splitext(os.path.basename(filepath))[0]
    header_row = [file_title]
    for job in job_numbers:
        header_row += ['Reg', 'O/T', 'Emp Tax', 'Memo 401k']

    top_row = ['Job Number']
    for job in job_numbers:
        top_row += [job, '', '', '']

    # Totals rows
    pay_total_row = ['pay total']
    grand_total_row = ['TOTAL']

    grand_sum_total_pay = 0
    grand_sum_total_tax = 0
    
    total_pay_per_job = {}
    total_emp_tax_per_job = {}
    total_memo_401k_per_job = {}

    for job in job_numbers:
        reg_col = pd.to_numeric(output_df[(job, 'Reg')], errors='coerce')
        ot_col = pd.to_numeric(output_df[(job, 'O/T')], errors='coerce')
        emp_tax_col = pd.to_numeric(output_df[(job, 'Emp Tax')], errors='coerce')
        memo_401k_col = pd.to_numeric(output_df[(job, 'Memo 401k')], errors='coerce')

        reg_total = reg_col.sum(skipna=True)
        ot_total = ot_col.sum(skipna=True)
        emp_tax_total = emp_tax_col.sum(skipna=True)
        memo_401k_total = memo_401k_col.sum(skipna=True)

        grand_sum_total_pay += reg_total + ot_total
        grand_sum_total_tax += emp_tax_total + memo_401k_total
        
        total_pay_per_job[job] = reg_total + ot_total
        total_emp_tax_per_job[job] = emp_tax_total
        total_memo_401k_per_job[job] = memo_401k_total

        pay_total_row.extend([round(reg_total, 2), round(ot_total, 2), round(emp_tax_total, 2), round(memo_401k_total, 2)])
        grand_total_row.extend([round(reg_total + ot_total + emp_tax_total + memo_401k_total, 2), '', '', ''])

    # Append final values to each row ---- .... not really needed and kinda misleading idk. We're reporting elsewhere (summary table) 
    # pay_total_row.append(round(grand_sum_total_pay, 2))
    # grand_total_row.append(round(grand_sum_total_pay + grand_sum_total_tax, 2))


    # Combine everything
    final_data = [top_row, header_row] + output_df.values.tolist() + [
        pay_total_row,
        grand_total_row
    ]

    final_df = pd.DataFrame(final_data)
    return final_df, grand_sum_total_pay, total_bonus, total_vacation, un_coded_pay, gross_total, total_pay_per_job, total_emp_tax_per_job, total_memo_401k_per_job, errors


def extract_job_costing_from_raw_excel(payroll_filepath, tax_filepath, output_file):
    errors = []
    
    # Process the employer tax file
    tax_data, errors_tax = calculate_tax(tax_filepath)
    if errors_tax:
        errors.extend(errors_tax)
        print("Errors in tax file processing:", errors_tax)
    
    # Process the payroll file
    payroll_df, total_pay, total_bonus, total_vacation, un_coded_pay, gross_total, total_pay_per_job, total_emp_tax_per_job, total_memo_401k_per_job, errors_payroll = process_payroll_file(payroll_filepath, tax_data)
    if errors_payroll:
        errors.extend(errors_payroll)
        print("Errors in payroll file processing:", errors_payroll)
    
    # Get all unique job numbers from both dictionaries
    all_jobs = sorted(set(total_pay_per_job.keys()).union(total_emp_tax_per_job.keys()))

    # Create summary rows: Job Number, Pay, Emp Tax, Total Cost
    summary_rows = []
    for job in all_jobs:
        pay = round(total_pay_per_job.get(job, 0), 2)
        emp_tax = round(total_emp_tax_per_job.get(job, 0), 2)
        memo_401k = round(total_memo_401k_per_job.get(job, 0), 2) if total_memo_401k_per_job.get(job, 0) else 0.0
        total_cost = round(pay + emp_tax + memo_401k, 2)
        summary_rows.append([job, pay, emp_tax, memo_401k, total_cost])

    # Convert to DataFrame
    summary_df = pd.DataFrame(summary_rows)
    
    # === Add 3 empty rows and the summary block ===
    final_df = pd.concat([
        payroll_df,
        pd.DataFrame([[]]*3),  # 3 empty rows
        pd.DataFrame([
            [round(total_pay, 2), 'Total Pay (Reg + O/T)'],
            [round(total_bonus, 2), 'Total Bonus'],
            [round(total_vacation, 2), 'Total Vacation'],
            [round(un_coded_pay, 2), 'Total Uncoded Pay'],
            [round(total_pay + total_bonus + total_vacation + un_coded_pay, 2), 'Sheet Total Pay + Bonus + Vacation + Uncoded'],
            [round(gross_total, 2), 'Report Total (Gross)'],  # New 5th summary row for gross pay
            [abs(round(gross_total - (total_pay + total_bonus + total_vacation + un_coded_pay), 2)), 'diff'],  # New 6th summary row for gross pay minus bonus
            
        ]),
        pd.DataFrame([[]]*3),  # 3 empty rows
        pd.DataFrame([['Job Number', 'Pay (Reg + O/T)', 'Emp Tax', 'Memo 401k', 'Total Job Cost']]),  # Header for the summary block        
        summary_df
        # # go through total pay per job, and add them to the end of the dataframe
        # pd.DataFrame([[job, round(total_pay_per_job[job], 2)] for job in total_pay_per_job.keys(), ]),
    ], ignore_index=True)

    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, header=False)

    # Set first column width to 165 pixels (~22.5 Excel width units)
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active
    ws.column_dimensions['A'].width = 25  # 1 Excel width unit ≈ 7 pixels
    
    # Set Memo 401k columns to 82px (~11.7 Excel width units, since 1 unit ≈ 7 pixels)
    memo_401k_width = 82 / 7  # ≈ 11.7

    # Find all columns with 'Memo 401k' in the header row (second row in your file)
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    for idx, col_name in enumerate(header_row, 1):
        if col_name == 'Memo 401k':
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = memo_401k_width
    
    wb.save(output_file)
    apply_formatting_to_excel(output_file)

    return output_file, errors

# Example:
# extract_job_costing_from_raw_excel("payroll/payroll_test_12.xls", "tax/tax_test_12.xlsx", "formatted_output_12.xlsx")
# extract_job_costing_from_raw_excel("payroll/payroll_test_16.xls", "tax/tax_test_16.xlsx", "formatted_output_16.xlsx")
# extract_job_costing_from_raw_excel("payroll/payroll_test_18.xls", "tax/tax_test_18.xlsx", "formatted_output_18.xlsx")
# extract_job_costing_from_raw_excel("payroll/payroll_test_20.xls", "tax/tax_test_20.xlsx", "formatted_output_20.xlsx")
