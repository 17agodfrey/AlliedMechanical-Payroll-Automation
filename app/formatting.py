from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def apply_formatting_to_excel(output_file):
    wb = load_workbook(output_file)
    ws = wb.active

    sea_green = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    blue_green = PatternFill(start_color="66CDAA", end_color="66CDAA", fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column

    # Format Reg/O/T header row (row 2), skip col 1
    for col in range(2, max_col + 1):
        ws.cell(row=2, column=col).fill = sea_green

    # Find the row that contains "TOTAL" in column 1
    total_row = None
    for row in range(3, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and isinstance(cell_value, str) and "TOTAL" in cell_value.upper():
            total_row = row
            break

    if total_row is None:
        # If no TOTAL row found, assume last row is total row
        total_row = max_row

    # Format Reg columns (every other col starting at 2), from row 3 up to row before TOTAL row
    for row in range(3, total_row):
        for col in range(2, max_col + 1, 2):  # Reg columns (every other column starting at 2)
            ws.cell(row=row, column=col).fill = light_blue

    # Format the TOTAL row only (row = total_row)
    for col in range(2, max_col + 1):
        ws.cell(row=total_row, column=col).fill = blue_green

    wb.save(output_file)
